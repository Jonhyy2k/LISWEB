# You need to login to the Bloomberg Terminal for the script to work!
# Run it using the arrow on the top right.
# Enter the stock ticker while specifying the country, for example Apple Inc. type AAPL US or 000660 KS; [TICKER] [COUNTRY] 
# All data is given in USD

import blpapi
import openpyxl
import shutil
import os
import time
from datetime import datetime

# Currency To USD (maybe)
overrides = request.getElement("overrides")
override = overrides.appendElement()
override.setElement("fieldId", "EQY_FUND_CRNCY")
override.setElement("value", "USD")


def setup_bloomberg_session(ticker_symbol):
    """Initialize Bloomberg API session with detailed logging."""
    options = blpapi.SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = blpapi.Session(options)
    
    print(f"[INFO] Attempting to connect to Bloomberg for {ticker_symbol}...")
    if not session.start():
        print("[WARNING] Failed to start Bloomberg session. Ensure Bloomberg Terminal is running.")
        return None
    if not session.openService("//blp/refdata"):
        print("[WARNING] Failed to open Bloomberg reference data service.")
        session.stop()
        return None
    print("[INFO] Bloomberg session started successfully.")
    return session

def fetch_bloomberg_data(session, ticker, fields, field_to_name_map, start_year=2014, end_year=2024, timeout=30):
    """Fetch historical data from Bloomberg with timeout and error handling."""
    if not fields:
        print("[INFO] No fields to fetch in this batch.")
        return {}
        
    if len(fields) > 25:
        raise ValueError(f"Too many fields ({len(fields)}). Bloomberg API limit is 25 fields per request.")
    
    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("HistoricalDataRequest")
    security = f"{ticker} Equity" 
    request.getElement("securities").appendValue(security)
    for field in fields:
        request.getElement("fields").appendValue(field)
    request.set("periodicitySelection", "YEARLY")
    request.set("startDate", f"{start_year}0101")
    request.set("endDate", f"{end_year}1231")
    
    print(f"[DEBUG] Sending request for {security} with fields: {fields}")
    session.sendRequest(request)
    
    data = {field: {} for field in fields}
    invalid_fields = []
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        event = session.nextEvent(500)
        if event.eventType() == blpapi.Event.TIMEOUT:
            print(f"[DEBUG] Bloomberg event timeout for {security}. Continuing to wait...")
            continue

        if event.eventType() in [blpapi.Event.RESPONSE, blpapi.Event.PARTIAL_RESPONSE]:
            for msg in event:
                print(f"[DEBUG] Raw Bloomberg response message: {msg}")  # Log raw response for debugging
                if msg.hasElement("responseError"):
                    error = msg.getElement("responseError")
                    error_message = error.getElement("message").getValue()
                    print(f"[ERROR] Bloomberg API error for {security}: {error_message}")
                    continue

                # FIXED SECTION: Handle securityData differently
                if not msg.hasElement("securityData"):
                    print(f"[WARNING] No securityData element in response for {security}.")
                    continue
                
                # securityData is not an array but a complex element - handle differently
                security_data = msg.getElement("securityData")
                
                # Handle field exceptions
                if security_data.hasElement("fieldExceptions"):
                    field_exceptions = security_data.getElement("fieldExceptions")
                    for j in range(field_exceptions.numValues()):
                        field_error = field_exceptions.getValue(j)
                        invalid_field_id = field_error.getElement("fieldId").getValueAsString()
                        error_info = field_error.getElement("errorInfo").getElement("message").getValueAsString()
                        field_name_display = field_to_name_map.get(invalid_field_id, "Unknown Field")
                        print(f"[WARNING] Invalid Bloomberg field: '{invalid_field_id}' (mapped to '{field_name_display}') for {security}. Error: {error_info}")
                        if invalid_field_id not in invalid_fields:
                            invalid_fields.append(invalid_field_id)
                
                # Check if fieldData exists 
                if not security_data.hasElement("fieldData"):
                    print(f"[WARNING] No fieldData element in securityData for {security}.")
                    continue

                # Process the fieldData array
                field_data_array = security_data.getElement("fieldData")
                print(f"[DEBUG] Number of fieldData entries: {field_data_array.numValues()}")
                
                for k in range(field_data_array.numValues()):
                    datum = field_data_array.getValue(k)
                    # Extract the date and convert to year
                    date_obj = datum.getElement("date").getValueAsDatetime()
                    year = date_obj.year
                    print(f"[DEBUG] Processing data for year {year}: {datum}")
                    
                    # Loop through each field and extract values
                    for field_id in fields:
                        if field_id in invalid_fields:
                            data[field_id][year] = "N/A (Invalid Field)"
                            continue
                        if datum.hasElement(field_id):
                            try:
                                value = datum.getElement(field_id).getValueAsFloat()
                                data[field_id][year] = value
                                print(f"[DEBUG] Fetched {field_id} for {year}: {value}")
                            except blpapi.exception.ElementErrorException:
                                try:
                                    value_str = datum.getElement(field_id).getValueAsString()
                                    data[field_id][year] = value_str
                                    print(f"[DEBUG] Field {field_id} for year {year} for {security} is not a float, stored as string: {value_str}")
                                except Exception as e_str:
                                    print(f"[WARNING] Could not get value for field {field_id} for year {year} for {security}: {e_str}")
                                    data[field_id][year] = "N/A (Error)"
                        else:
                            if year not in data[field_id]:
                                data[field_id][year] = None
                                print(f"[DEBUG] No data for {field_id} in {year} for {security}")

        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg in event:
                if msg.messageType() == blpapi.Name("SessionTerminated"):
                    print("[ERROR] Bloomberg session terminated unexpectedly.")
                    return None

        if event.eventType() == blpapi.Event.RESPONSE:
            print(f"[INFO] Received final response for batch for {security}.")
            break
    
    if time.time() - start_time >= timeout and event.eventType() != blpapi.Event.RESPONSE:
        print(f"[WARNING] Timed out waiting for complete response for {security} after {timeout}s.")

    if not any(data[field] for field in data if data[field]):
        print(f"[WARNING] No data received for any requested field for {ticker} in this batch.")
    
    if invalid_fields:
        print(f"[INFO] Bloomberg fields skipped or marked N/A due to invalidity for {security}: {invalid_fields}")
    
    print(f"[DEBUG] Final fetched data: {data}")
    return data

def calculate_derived_metrics(data, start_year=2014, end_year=2024):
    """Calculate derived metrics like DSO."""
    derived = {
        "DSO": {}
    }
    
    def get_val(source_field, year, default=0):
        val = data.get(source_field, {}).get(year)
        if isinstance(val, (int, float)):
            return val
        return default

    for year in range(start_year, end_year + 1):
        revenue = get_val("SALES_REV_TURN", year)
        ar = get_val("BS_ACCT_NOTE_RCV", year)
        derived["DSO"][year] = (ar / revenue * 365) if revenue else 0

    return derived

field_map = {
    # Income Statement (IS)
    "Revenue (Sales)": {"source": "BDH", "field": "SALES_REV_TURN", "statement": "IS"},
    "COGS (Cost of Goods Sold)": {"source": "BDH", "field": "IS_COG_AND_SERVICES_SOLD", "statement": "IS"},
    "Gross Profit": {"source": "BDH", "field": "GROSS_PROFIT", "statement": "IS"},
    "SG&A (Selling, General & Administrative)": {"source": "BDH", "field": "IS_SGA_EXPENSE", "statement": "IS"},
    "R&D (Research & Development)": {"source": "BDH", "field": "IS_OPERATING_EXPENSES_RD", "statement": "IS"},
    "Other Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_OPER_INC", "statement": "IS"},
    "EBITDA": {"source": "BDH", "field": "EBITDA", "statement": "IS"},
    "D&A (Depreciation & Amortization)": {"source": "BDH", "field": "ARDR_DEPRECIATION_AMORTIZATION", "statement": "IS"},
    "Depreciation Expense": {"source": "BDH", "field": "ARDR_DEPRECIATION_EXP", "statement": "IS"},
    "Amortization Expense": {"source": "BDH", "field": "ARDR_AMORT_EXP", "statement": "IS"},
    "Operating Income (EBIT)": {"source": "BDH", "field": "IS_OPER_INC", "statement": "IS"},
    "Net Interest Expense (Income)": {"source": "BDH", "field": "IS_NET_INTEREST_EXPENSE", "statement": "IS"},
    "Interest Expense": {"source": "BDH", "field": "IS_INT_EXPENSE", "statement": "IS"},
    "Interest Income": {"source": "BDH", "field": "IS_INT_INC", "statement": "IS"},
    "FX (Gain) Loss": {"source": "BDH", "field": "IS_FOREIGN_EXCH_LOSS", "statement": "IS"},
    "Other Non-Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_NON_OPERATING_INC_LOSS", "statement": "IS"},
    "Pre-Tax Income (EBT)": {"source": "BDH", "field": "PRETAX_INC", "statement": "IS"},
    "Tax Expense (Benefits)": {"source": "BDH", "field": "IS_INC_TAX_EXP", "statement": "IS"},
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "IS"},
    "EPS Basic": {"source": "BDH", "field": "BASIC_EPS", "statement": "IS"},
    "EPS Diluted": {"source": "BDH", "field": "DILUTED_EPS", "statement": "IS"},
    "Basic Weighted Average Shares": {"source": "BDH", "field": "IS_AVG_NUM_SH_FOR_EPS", "statement": "IS"},
    "Diluted Weighted Average Shares": {"source": "BDH", "field": "IS_SH_FOR_DILUTED_EPS", "statement": "IS"},

    # Balance Sheet (BS)
    #"Cash & Cash Equivalents & ST Investments": {"source": "BDH", "field": "CASH_CASH_EQTY_STI_DETAILED", "statement": "BS"},
    "Cash & Cash Equivalents": {"source": "BDH", "field": "BS_CASH_NEAR_CASH_ITEM", "statement": "BS"},
    "Short-Term Investments": {"source": "BDH", "field": "BS_MKT_SEC_OTHER_ST_INVEST", "statement": "BS"},
    "Accounts Receivable": {"source": "BDH", "field": "BS_ACCT_NOTE_RCV", "statement": "BS"},
    "Inventory": {"source": "BDH", "field": "BS_INVENTORIES", "statement": "BS"},
    #"Prepaid Expenses and Other Current Assets": {"source": "BDH", "field": "OTH_CUR_ASSETS", "statement": "BS"},
    "Current Assets": {"source": "BDH", "field": "BS_CUR_ASSET_REPORT", "statement": "BS"},
    #"Net PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "NET_PPE", "statement": "BS"},
    "Gross PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "BS_GROSS_FIX_ASSET", "statement": "BS"},
    "Accumulated Depreciation": {"source": "BDH", "field": "BS_ACCUM_DEPR", "statement": "BS"},
    #"Right-of-Use Assets": {"source": "BDH", "field": "OPER_LEASE_ASSETS", "statement": "BS"},
    "Intangibles": {"source": "BDH", "field": "BS_DISCLOSED_INTANGIBLES", "statement": "BS"},
    "Goodwill": {"source": "BDH", "field": "BS_GOODWILL", "statement": "BS"},
    #"Intangibles excl. Goodwill": {"source": "BDH", "field": "NET_OTHER_INTAN_ASSETS", "statement": "BS"},
    #"Other Non-Current Assets": {"source": "BDH", "field": "OTH_NON_CUR_ASSETS", "statement": "BS"},
    "Non-Current Assets": {"source": "BDH", "field": "BS_TOT_NON_CUR_ASSET", "statement": "BS"},
    #"Total Assets": {"source": "BDH", "field": "TOT_ASSETS", "statement": "BS"},
    "Accounts Payable": {"source": "BDH", "field": "BS_ACCT_PAYABLE", "statement": "BS"},
    #"Short-Term Debt": {"source": "BDH", "field": "ST_DEBT", "statement": "BS"},
    "Short-Term Borrowings": {"source": "BDH", "field": "SHORT_TERM_DEBT_DETAILED", "statement": "BS"},
    "Current Portion of Lease Liabilities": {"source": "BDH", "field": "ST_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    #"Accrued Expenses and Other Current Liabilities": {"source": "BDH", "field": "OTH_CUR_LIAB", "statement": "BS"},
    "Current Liabilities": {"source": "BDH", "field": "BS_CUR_LIAB", "statement": "BS"},
    "Long-Term Borrowings": {"source": "BDH", "field": "LONG_TERM_BORROWINGS_DETAILED", "statement": "BS"},
    "Long-Term Operating Lease Liabilities": {"source": "BDH", "field": "LT_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Non-Current Liabilities": {"source": "BDH", "field": "NON_CUR_LIAB", "statement": "BS"},
    "Current Liabilities": {"source": "BDH", "field": "BS_CUR_LIAB", "statement": "BS"},
    "Non-Controlling Interest": {"source": "BDH", "field": "MINORITY_NONCONTROLLING_INTEREST", "statement": "BS"},

    # Cash Flow Statement (CF) - Fields are typically changes or actual cash flows
    "(Increase) Decrease in Accounts Receivable": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Inventories": {"source": "BDH", "field": "CF_CHANGE_IN_INVENTORIES", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Pre-paid expeses and Other CA": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "Increase (Decrease) in Accounts Payable": {"source": "BDH", "field": "CF_CHANGE_IN_ACCOUNTS_PAYABLE", "statement": "CF", "section": "Operating"},
    "Increase (Decrease) in Accrued Revenues and Other CL": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "Stock Based Compensation": {"source": "BDH", "field": "CF_STOCK_BASED_COMPENSATION", "statement": "CF", "section": "Operating"},
    #"Other Operating Adjustments": {"source": "BDH", "field": "CF_OTHER_OPERATING_ACT", "statement": "CF", "section": "Operating"}, # Catch-all
    "Operating Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_OPER", "statement": "CF", "section": "Operating"},
    #"Net Capex": {"source": "BDH", "field": "ARD_CAPITAL_EXPENDITURES", "statement": "CF", "section": "Investing"}, # CAPITAL_EXPEND is common
    "Acquisition of Fixed & Intangibles": {"source": "BDH", "field": "ACQUIS_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Disposal of Fixed & Intangibles": {"source": "BDH", "field": "DISPOSAL_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Acquisitions": {"source": "BDH", "field": "CF_CASH_FOR_ACQUIS_SUBSIDIARIES", "statement": "CF", "section": "Investing"},
    "Divestitures": {"source": "BDH", "field": "CF_CASH_FOR_DIVESTITURES", "statement": "CF", "section": "Investing"},
    "Increase in LT Investment": {"source": "BDH", "field": "CF_INCR_INVEST", "statement": "CF", "section": "Investing"},
    "Decrease in LT Investment": {"source": "BDH", "field": "CF_DECR_INVEST", "statement": "CF", "section": "Investing"},
    #"Other Investing Inflows (Outflows)": {"source": "BDH", "field": "OTHER_INVESTING_ACT_DETAILED", "statement": "CF", "section": "Investing"},
    "Investing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_INV_ACT", "statement": "CF", "section": "Investing"},
    #"Lease Payments": {"source": "BDH", "field": "CF_LEASE_PAYMENTS", "statement": "CF", "section": "Financing"}, # Principal portion
    "Debt Borrowing": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PROCEEDS", "statement": "CF", "section": "Financing"},
    "Debt Repayment": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PAYMENT", "statement": "CF", "section": "Financing"},
    "Dividends": {"source": "BDH", "field": "CF_DVD_PAID", "statement": "CF", "section": "Financing"},
    "Increase (Repurchase) of Shares": {"source": "BDH", "field": "PROC_FR_REPURCH_EQTY_DETAILED", "statement": "CF", "section": "Financing"}, # Net issuance/repurchase
    #"Other Financing Inflows (Outflows)": {"source": "BDH", "field": "OTHER_FIN_AND_DEC_CAP", "statement": "CF", "section": "Financing"},
    "Financing Cash Flow": {"source": "BDH", "field": "CFF_ACTIVITIES_DETAILED", "statement": "CF", "section": "Financing"}, # CASH_FLOW_FDS_FIN_ACT common
    "Effect of Foreign Exchange": {"source": "BDH", "field": "CF_EFFECT_FOREIGN_EXCHANGES", "statement": "CF", "section": "All"}, # Applies to overall CF reconciliation

    # Additional Fields (BS)
    "Market Capitalization": {"source": "BDH", "field": "CUR_MKT_CAP", "statement": "BS"},
    "Total Debt": {"source": "BDH", "field": "SHORT_AND_LONG_TERM_DEBT", "statement": "BS"},
    "Preferred Stock": {"source": "BDH", "field": "PFD_EQTY_HYBRID_CAPITAL", "statement": "BS"},
    "Non-Controlling Interest": {"source": "BDH", "field": "MINORITY_NONCONTROLLING_INTEREST", "statement": "BS"},
    "Enterprise Value": {"source": "BDH", "field": "ENTERPRISE_VALUE", "statement": "BS"},
    "Total Borrowings": {"source": "BDH", "field": "TOT_BORROWINGS", "statement": "BS"},
    "Total Leases": {"source": "BDH", "field": "TOT_LEASE_LIAB", "statement": "BS"},
    "Net Debt": {"source": "BDH", "field": "NET_DEBT", "statement": "BS"},
    "Effective Tax Rate": {"source": "BDH", "field": "EFF_TAX_RATE", "statement": "BS"},

    # Derived Metrics
    "Changes in Net Working Capital": {"source": "derived", "field": "Changes in Net Working Capital", "statement": "BS"},
    "DSO": {"source": "derived", "field": "DSO", "statement": "IS"},
    "DIH": {"source": "derived", "field": "DIH", "statement": "BS"},
    "DPO": {"source": "derived", "field": "DPO", "statement": "BS"},
    "Net Cash from Investments & Acquisitions": {"source": "derived", "field": "Net Cash from Investments & Acquisitions", "statement": "CF", "section": "Investing"},
    "Increase (Decrease) in Other": {"source": "derived", "field": "Increase (Decrease) in Other", "statement": "CF", "section": "Operating"},

}

# Manual cell mapping for 2014 data
field_cell_map = {

    # Income Statement (IS)
    "Revenue (Sales)": "G6",
    "COGS (Cost of Goods Sold)": "G7",
    "Gross Profit": "G8",
    "SG&A (Selling, General & Administrative)": "G9",
    "R&D (Research & Development)": "G10",
    "Other Operating (Income) Expenses": "G11",
    "EBITDA": "G12",
    "D&A (Depreciation & Amortization)": "G13",
    "Depreciation Expense": "G14",
    "Amortization Expense": "G15",
    "Operating Income (EBIT)": "G16",
    "Net Interest Expense (Income)": "G17",
    "Interest Expense": "G18",
    "Interest Income": "G19",
    "FX (Gain) Loss": "G20",
    "Other Non-Operating (Income) Expenses": "G21",
    "Pre-Tax Income (EBT)": "G22",
    "Tax Expense (Benefits)": "G23",
    "Net Income": "G24",
    "EPS Basic": "G25",
    "EPS Diluted": "G26",
    "Basic Weighted Average Shares": "G27",
    "Diluted Weighted Average Shares": "G28",

    # Balance Sheet (BS)
    #"Cash & Cash Equivalents & ST Investments": "G32",
    "Cash & Cash Equivalents": "G33",
    "Short-Term Investments": "G34",
    "Accounts Receivable": "G35",
    "Inventory": "G36",
    #"Prepaid Expenses and Other Current Assets": "G37",
    "Current Assets": "G38",
    #"Net PP&E (Property, Plant and Equipment)": "G39",
    "Gross PP&E (Property, Plant and Equipment)": "G40",
    "Accumulated Depreciation": "G41",
    #"Right-of-Use Assets": "G42",
    "Intangibles": "G43",
    "Goodwill": "G44",
    #"Intangibles excl. Goodwill": "G45",
    #"Other Non-Current Assets": "G46",
    "Non-Current Assets": "G47",
    #"Total Assets": "G48",
    "Accounts Payable": "G49",
    #"Short-Term Debt": "G50",
    "Short-Term Borrowings": "G51",
    "Current Portion of Lease Liabilities": "G52",
    #"Accrued Expenses and Other Current Liabilities": "G52",
    "Current Liabilities": "G54",
    #"Long-Term Debt": "G54",
    "Long-Term Borrowings": "G56",
    "Long-Term Operating Lease Liabilities": "G57",
    "Non-Current Liabilities": "G59",
    "Non-Controlling Interest": "G62",

    # Cash Flow Statement (CF)
    #"Net Income": "G66",
    "(Increase) Decrease in Accounts Receivable": "G69",
    "(Increase) Decrease in Inventories": "G70",
    "(Increase) Decrease in Pre-paid expeses and Other CA": "G71",
    "Increase (Decrease) in Accounts Payable": "G72",
    "Increase (Decrease) in Accrued Revenues and Other CL": "G73", 
    "Stock Based Compensation": "G74",
    #"Other Operating Adjustments": "G70",
    "Operating Cash Flow": "G76",
    #"Increase (Decrease) in Other": "G72",
    #"Net Capex": "G73",
    "Acquisition of Fixed & Intangibles": "G78",
    "Disposal of Fixed & Intangibles": "G79",
    "Acquisitions": "G81",
    "Divestitures": "G82",
    "Increase in LT Investment": "G83",
    "Decrease in LT Investment": "G84",
    #"Other Investing Inflows (Outflows)": "G80",
    "Investing Cash Flow": "G86",
    #"Net Cash from Investments & Acquisitions": "G82",
    #"Lease Payments": "G89",
    "Debt Borrowing": "G87",
    "Debt Repayment": "G88",
    "Dividends": "G90",
    "Increase (Repurchase) of Shares": "G91",
    #"Other Financing Inflows (Outflows)": "G88",
    "Financing Cash Flow": "G93",
    "Effect of Foreign Exchange": "G94",

    # Additional Fields (BS)
    "Market Capitalization": "G99",
    "Total Debt": "G101",
    "Preferred Stock": "G102",
    "Non-Controlling Interest": "G103",
    "Enterprise Value": "G104",
    #"Total Borrowings": "G96",
    #"Total Leases": "G116",
    #"Net Debt": "G98",
    #"Effective Tax Rate": "G99",
    # Other Derived Metrics
    #"Changes in Net Working Capital": "G100",
    #"DSO": "G101",
    #"DIH": "G102",
    #"DPO": "G103"
}

def filter_field_map_for_task(task_name):
    """Filters the field_map for a specific task (e.g., 'IS', 'BS')."""
    statement_code, cf_section = task_name.split("_") if "_" in task_name else (task_name, None)
    
    allowed_statements = ["IS", "BS", "CF"]
    if statement_code not in allowed_statements:
        raise ValueError(f"Invalid statement code '{statement_code}'. Must be one of {allowed_statements}.")
    
    task_specific_fields = {}
    for name, config in field_map.items():
        if config["statement"] == statement_code:
            task_specific_fields[name] = config
    
    required_bdh_for_derived = set()
    for name, config in task_specific_fields.items():
        if config["source"] == "derived":
            if name == "DSO":
                required_bdh_for_derived.update(["BS_ACCT_NOTE_RCV", "SALES_REV_TURN"])
    
    for bdh_field_code in required_bdh_for_derived:
        found = False
        for name, config in task_specific_fields.items():
            if config.get("field") == bdh_field_code and config.get("source") == "BDH":
                found = True
                break
        if not found:
            for name, global_config in field_map.items():
                if global_config.get("field") == bdh_field_code and global_config.get("source") == "BDH":
                    task_specific_fields[f"__dep_{name}"] = global_config
                    break
    return task_specific_fields

def batch_fields(fields_to_fetch, batch_size=25):
    """Split fields into batches of batch_size or fewer."""
    unique_fields = sorted(list(set(fields_to_fetch)))
    return [unique_fields[i:i + batch_size] for i in range(0, len(unique_fields), batch_size)]

def get_column_letter_from_index(col_index):
    """Convert 1-based column index to letter (e.g., 7 -> G)."""
    return openpyxl.utils.get_column_letter(col_index)

def get_target_cells_for_years(base_cell_ref, num_years):
    """Get list of cell references for a row, for num_years, starting from base_cell_ref."""
    try:
        col_str = "".join(filter(str.isalpha, base_cell_ref))
        row_num = int("".join(filter(str.isdigit, base_cell_ref)))
        start_col_idx = openpyxl.utils.column_index_from_string(col_str)
        
        target_cells = []
        for i in range(num_years):
            target_col_letter = get_column_letter_from_index(start_col_idx + i)
            target_cells.append(f"{target_col_letter}{row_num}")
        return target_cells
    except ValueError as e:
        print(f"[ERROR] Invalid base cell reference '{base_cell_ref}': {e}")
        raise

def populate_valuation_model(template_path, output_path, ticker_symbol):
    """Populate the 'Inputs' sheet with data for all statements in a single file."""
    if not os.path.exists(template_path):
        print(f"[ERROR] Template file '{template_path}' not found.")
        raise FileNotFoundError(f"Template file {template_path} not found.")

    shutil.copy(template_path, output_path)
    print(f"[INFO] Copied template '{template_path}' to output file '{output_path}'.")

    wb = openpyxl.load_workbook(output_path)
    if "Inputs" not in wb.sheetnames:
        print("[ERROR] 'Inputs' sheet not found in the workbook.")
        raise ValueError("'Inputs' sheet not found in the template file.")
    ws = wb["Inputs"]

    tasks_to_process = ["IS", "BS", "CF"]
    data_years = list(range(2014, 2025))

    all_fetched_bdh_data = {}
    global_bberg_code_to_name_map = {
        config["field"]: name
        for name, config in field_map.items()
        if config["source"] == "BDH" and "field" in config
    }

    print(f"\n[PHASE] Starting data fetching for ticker: {ticker_symbol}")
    for task_name in tasks_to_process:
        print(f"\n  [TASK] Processing data collection for: {task_name}")

        current_task_field_configs = filter_field_map_for_task(task_name)
        bdh_fields_for_this_task = [
            config["field"]
            for name, config in current_task_field_configs.items()
            if config["source"] == "BDH" and "field" in config
        ]

        if not bdh_fields_for_this_task:
            print(f"    [INFO] No Bloomberg (BDH) fields to fetch for task '{task_name}'. Skipping.")
            continue

        field_batches = batch_fields(bdh_fields_for_this_task, batch_size=25)

        for batch_idx, current_batch_fields in enumerate(field_batches):
            session = None
            try:
                session = setup_bloomberg_session(ticker_symbol)
                if not session:
                    continue

                # Apply USD override
                ref_data_service = session.getService("//blp/refdata")
                request = ref_data_service.createRequest("HistoricalDataRequest")
                request.getElement("securities").appendValue(ticker_symbol)
                for field in current_batch_fields:
                    request.getElement("fields").appendValue(field)
                request.set("periodicitySelection", "YEARLY")
                request.set("startDate", "20140101")
                request.set("endDate", "20241231")

                overrides = request.getElement("overrides")
                override = overrides.appendElement()
                override.setElement("fieldId", "EQY_FUND_CRNCY")
                override.setElement("value", "USD")

                print(f"[DEBUG] Sending USD override request for batch {batch_idx + 1}")
                session.sendRequest(request)

                batch_data_fetched = fetch_bloomberg_data(
                    session, ticker_symbol, current_batch_fields,
                    global_bberg_code_to_name_map, start_year=2014, end_year=2024
                )

                if batch_data_fetched:
                    for field_code, yearly_data in batch_data_fetched.items():
                        if field_code not in all_fetched_bdh_data:
                            all_fetched_bdh_data[field_code] = {}
                        for year, value in yearly_data.items():
                            all_fetched_bdh_data[field_code][year] = value

            finally:
                if session:
                    session.stop()

    print(f"\n[PHASE] Completed all data fetching.")

    print(f"\n[PHASE] Calculating derived metrics...")
    all_derived_data = calculate_derived_metrics(all_fetched_bdh_data, start_year=2014, end_year=2024)
    print("[INFO] Derived metrics calculated.")

    print(f"\n[PHASE] Writing all data to Excel sheet '{ws.title}'...")

    for item_name, config in field_map.items():
        if item_name.startswith("__dep_"):
            continue

        base_cell_ref = field_cell_map.get(item_name)
        if not base_cell_ref:
            continue

        target_cells_for_item = get_target_cells_for_years(base_cell_ref, len(data_years))

        if config["source"] == "BDH":
            bberg_field_code = config["field"]
            data_source_for_item = all_fetched_bdh_data.get(bberg_field_code, {})
            for i, year in enumerate(data_years):
                cell_ref = target_cells_for_item[i]
                raw_value = data_source_for_item.get(year)

                if isinstance(raw_value, (int, float)):
                    ws[cell_ref] = raw_value
                    ws[cell_ref].number_format = "#,##0.000"
                elif isinstance(raw_value, str) and "N/A" in raw_value:
                    ws[cell_ref] = raw_value
                else:
                    ws[cell_ref] = 0

        elif config["source"] == "derived":
            data_source_for_item = all_derived_data.get(config["field"], {})
            for i, year in enumerate(data_years):
                cell_ref = target_cells_for_item[i]
                value = data_source_for_item.get(year)
                if isinstance(value, (int, float)):
                    ws[cell_ref] = value
                    if "DSO" in item_name:
                        ws[cell_ref].number_format = "0.0"
                    else:
                        ws[cell_ref].number_format = "#,##0.000"
                else:
                    ws[cell_ref] = 0

    wb.save(output_path)
    print(f"\n[SUCCESS] Valuation model populated and saved to '{output_path}'")

if __name__ == "__main__":
    print("-" * 70)
    print("Bloomberg Data to Excel Valuation Model Populator")
    print("-" * 70)
    print("This script fetches financial data for IS, BS, and CF statements,")
    print("processes it in batches, calculates derived metrics, and populates")
    print("a single Excel template.")
    print("Ensure Bloomberg Terminal is running and blpapi is correctly configured.")
    print("-" * 70)

    excel_template_path = "LIS_Valuation_Empty.xlsx"
    output_directory = "."
    
    ticker_input = ""
    while not ticker_input:
        ticker_input = input("Enter the Ticker Symbol (e.g., AAPL US Equity for Apple Inc.): ").strip().upper()
        if not ticker_input:
            print("[VALIDATION] Ticker symbol cannot be empty. Please try again.")

    output_file_name = f"{ticker_input}_Valuation_Model_{datetime.now().strftime('%Y%m%d')}.xlsx"
    final_output_path = os.path.join(output_directory, output_file_name)
    
    print(f"\n[SETUP] Template: '{excel_template_path}'")
    print(f"[SETUP] Output will be: '{final_output_path}'")
    print(f"[SETUP] Ticker: '{ticker_input}'")
    
    try:
        print("\nStarting the data population process...\n")
        populate_valuation_model(
            template_path=excel_template_path,
            output_path=final_output_path,
            ticker_symbol=ticker_input
        )
        print("\nProcess completed successfully.")
    except Exception as e:
        print(f"[CRITICAL ERROR] An unexpected error occurred: {e}")
        import traceback
        print("\n--- Traceback ---")
        traceback.print_exc()
        print("--- End Traceback ---\n")
    finally:
        print("\nScript execution finished.")
